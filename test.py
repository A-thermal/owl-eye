#遍历寻找目标值位置
from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl import worksheet

def comparison(a,b):
    ib=0
    for ia in range(len(a)):
        if ord(a[ia:ia+1])-ord(b[ib:ib+1])==0:
            ib=ib+1
            if ib==len(b):
                print('a and b are equall')
        else:
            print('a and b are not equall')
            break

def find_false_in_sheet(sheet):
    for column in sheet.iter_cols():
        for cell2 in column:
            if cell2.value is not None:
                if type(cell2.value)==int:
                    info2 = ord(str(cell2.value))-ord('核能')#ord()仅支持一个字的，两个字就不得行了，此处设计一个，逐个字比对的函数
                    if info2 == 0:
                        print (cell2)
                        print (cell2.value)
                elif type(cell2.value)==str:
                # print cell2.value
                # print type(cell2.value)
                    info2 = ord(cell2.value)-ord('核能')
                    if info2 == 0:
                        print (cell2)
                        print (cell2.value)

def find_false_in_xlsx(file_name):
    print (file_name)
    wb = load_workbook(file_name)
    all_sheets = wb.get_sheet_names()
    print(all_sheets)

    for i in range(len(all_sheets)):
        sheet = wb.get_sheet_by_name(all_sheets[i])
        print (sheet.title + ': max_row: ' + str(sheet.max_row) + '  max_column: ' + str(sheet.max_column))
        find_false_in_sheet(sheet)



find_false_in_xlsx('f:/2018/test.xlsx')
